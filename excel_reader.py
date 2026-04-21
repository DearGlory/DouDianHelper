from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook


DEFAULT_ALIASES = [
    "订单号",
    "订单编号",
    "主订单号",
    "父订单号",
    "订单ID",
    "order_id",
    "orderid",
    "订单",
]


def _normalize_order_id(value) -> str:
    """Convert cell value to a clean order ID string.

    Handles: int, float (strips .0), scientific notation, str with whitespace.
    """
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    if re.match(r"^\d+\.?\d*[eE]\+?\d+$", s):
        try:
            return str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    if s.endswith(".0"):
        s = s[:-2]
    return s


class ExcelOrderReader:
    def __init__(self, excel_path: str, aliases: Iterable[str] | None = None):
        self.excel_path = Path(excel_path)
        self.aliases = [a.strip().lower() for a in (aliases or DEFAULT_ALIASES)]

    def _normalize_header(self, value) -> str:
        if value is None:
            return ""
        return str(value).strip().lower().replace(" ", "")

    def _detect_order_id_col(self, header_row: List[object]) -> int:
        normalized = [self._normalize_header(cell) for cell in header_row]
        for idx, name in enumerate(normalized):
            if not name:
                continue
            for alias in self.aliases:
                if alias.replace(" ", "") == name:
                    return idx
        for idx, name in enumerate(normalized):
            if "订单" in name and ("号" in name or "id" in name or "编号" in name):
                return idx
        raise ValueError("未识别到订单号列，请检查 Excel 表头或在 config 中补充 order_id_aliases")

    def read_order_ids(self) -> list[str]:
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {self.excel_path}")

        wb = load_workbook(self.excel_path, read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]

        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            wb.close()
            return []

        order_col = self._detect_order_id_col(list(header))
        seen: set[str] = set()
        results: list[str] = []

        for row in rows:
            if row is None or order_col >= len(row):
                continue
            if all(c is None for c in row):
                continue
            order_id = _normalize_order_id(row[order_col])
            if not order_id:
                continue
            if order_id in seen:
                continue
            seen.add(order_id)
            results.append(order_id)

        wb.close()
        return results

    def fetch_rows_by_order_ids(self, order_ids: Iterable[str]) -> tuple[list[object], list[list[object]]]:
        normalized_targets = {_normalize_order_id(order_id) for order_id in order_ids if _normalize_order_id(order_id)}
        wb = load_workbook(self.excel_path, read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            wb.close()
            return [], []
        order_col = self._detect_order_id_col(list(header))
        matched: list[list[object]] = []
        for row in rows:
            if row is None or order_col >= len(row):
                continue
            if all(c is None for c in row):
                continue
            order_id = _normalize_order_id(row[order_col])
            if order_id in normalized_targets:
                matched.append(list(row))
        wb.close()
        return list(header), matched

    def export_rows_with_results(self, output_path: str, header: list[object], rows: list[list[object]], results_by_order_id: dict[str, dict[str, str]]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "processed_orders"
        ws.append(list(header) + ["run_status", "run_reason"])
        order_col = self._detect_order_id_col(list(header))
        for row in rows:
            order_id = _normalize_order_id(row[order_col] if order_col < len(row) else "")
            result = results_by_order_id.get(order_id, {})
            ws.append(list(row) + [result.get("status", ""), result.get("reason", "")])
        wb.save(output_path)
        wb.close()

    def delete_order_rows_bulk(self, order_ids: Iterable[str]) -> int:
        normalized_targets = {_normalize_order_id(order_id) for order_id in order_ids if _normalize_order_id(order_id)}
        if not normalized_targets:
            return 0

        wb = load_workbook(self.excel_path, read_only=False)
        ws = wb[wb.sheetnames[0]]

        header = [cell.value for cell in ws[1]]
        order_col = self._detect_order_id_col(header)
        delete_count = 0

        next_row = 2
        max_row = ws.max_row
        while next_row <= max_row:
            cell_value = ws.cell(row=next_row, column=order_col + 1).value
            if _normalize_order_id(cell_value) in normalized_targets:
                ws.delete_rows(next_row)
                delete_count += 1
                max_row -= 1
                continue
            next_row += 1

        wb.save(self.excel_path)
        wb.close()
        return delete_count
